#!/usr/bin/env python3
"""
Синхронизация экзаменационных билетов с Google Sheets.

Источник: https://docs.google.com/spreadsheets/d/1D8ElnUF3_ucNCpl0kF3PcVGltF08uoJK/edit
          (файл «Экзаменационные билеты_app.xlsx», импортированный в Google Sheets)

Скрипт работает по тому же принципу, что и scripts/sync-projects.py
(раздел «Проекты»):

  1. Скачивает XLSX-экспорт напрямую из Google Sheets через export?format=xlsx.
     Google отдаёт файл без OAuth, если таблица доступна «у кого есть ссылка».
  2. Парсит 4 листа («4 разряд», «5 разряд», «6 разряд», «До 1000 В»).
  3. Сохраняет результат в data/exam-tickets.json.

Переменные окружения:
  EXAM_TICKETS_SPREADSHEET_ID — ID Google Sheets
      (по умолчанию 1D8ElnUF3_ucNCpl0kF3PcVGltF08uoJK)
  EXAM_TICKETS_GID — numeric ID листа (опционально; если задан, экспортирует
      конкретный лист через &gid=...). Если не задан — экспортируется вся книга.

Секреты НЕ требуются — таблица доступна «у кого есть ссылка»,
Google отдаёт XLSX через export?format=xlsx без OAuth.

Если нет интернета или API недоступен — используется уже существующий
data/exam-tickets.json как заглушка (PWA продолжает работать с последними
закоммиченными данными).

История источников:
- OneDrive (legacy) → Google Sheets (текущий)
"""

import json
import os
import sys
from pathlib import Path
from datetime import datetime, time as dtime, date as ddate

import requests
import openpyxl


# ============================================================
# Настройки Google Sheets
# ============================================================
DEFAULT_SPREADSHEET_ID = '1D8ElnUF3_ucNCpl0kF3PcVGltF08uoJK'

DOWNLOAD_DIR = Path('/tmp/exam_tickets_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'exam-tickets.json'


# ============================================================
# Конфигурация листов
# ============================================================
SHEETS_CONFIG = {
    "4 разряд": {"id": "tickets-4", "title": "Билеты на 4 разряд"},
    "5 разряд": {"id": "tickets-5", "title": "Билеты на 5 разряд"},
    "6 разряд": {"id": "tickets-6", "title": "Билеты на 6 разряд"},
    "До 1000 В": {"id": "tickets-1000v", "title": "Билеты до 1000 В"},
}


# ============================================================
# Стандартизация имён полей
# ============================================================
# Excel-таблица содержит русские заголовки с пробелами и спецсимволами
# (№ билета, № вопроса, Название литературы и т.д.).
# Для JSON используем латинские snake_case имена — это упрощает
# работу в JavaScript, устраняет проблемы с кодировками в escape-
# последовательностях (\u2116 для № и т.д.) и делает код читаемее.
#
# Фронтенд PWA (index.html) читает новые имена, но также поддерживает
# старые русские имена как fallback — это обеспечивает плавную
# миграцию, если где-то остался старый JSON.
FIELD_NAME_MAP = {
    "ID": "id",
    "№ билета": "ticket_number",
    "№билета": "ticket_number",        # вариант без пробела (для надёжности)
    "№ вопроса": "question_number",
    "№вопроса": "question_number",     # вариант без пробела
    "Вопрос": "question",
    "Ответ": "answer",
    "Image": "image_url",
    "Название литературы": "literature_name",
    "Файл": "file_url",
}


def log(msg):
    print(f'[exam-tickets] {msg}', flush=True)


# ============================================================
# Восстановление числового значения из «даты/времени»
# ============================================================
# В Google Sheets числовые колонки иногда имеют формат Date/Time.
# openpyxl с data_only=True возвращает datetime/time вместо числа.
# Решение: конвертируем datetime/time обратно в Excel serial number.
# Эпоха 1899-12-30 = 1900 date system (как в Google Sheets и Excel).
DATE_EPOCH = datetime(1899, 12, 30)

def datetime_to_serial(val):
    """Конвертирует datetime/time/date в Excel serial number (float).
    Возвращает None, если конвертация неприменима.
    """
    if isinstance(val, datetime):
        delta = val - DATE_EPOCH
        return round(delta.total_seconds() / 86400.0, 6)
    if isinstance(val, dtime):
        secs = val.hour * 3600 + val.minute * 60 + val.second + val.microsecond / 1e6
        return round(secs / 86400.0, 6)
    if isinstance(val, ddate):
        delta = datetime(val.year, val.month, val.day) - DATE_EPOCH
        return round(delta.total_seconds() / 86400.0, 6)
    return None


def format_serial_as_string(serial):
    """Форматирует serial number: int если целое, иначе trimmed float."""
    if abs(serial - round(serial)) < 1e-9:
        return str(int(round(serial)))
    return f'{serial:.4f}'.rstrip('0').rstrip('.')


# Числовые поля в таблице билетов (после нормализации имён).
# Если в Google Sheets эти колонки имеют формат Date/Time,
# конвертируем datetime/time обратно в serial number.
NUMERIC_FIELD_NAMES = {'id', 'ticket_number', 'question_number'}


def _normalize_field_name(raw_name: str) -> str:
    """Конвертирует русское имя поля из Excel в стандартное латинское.

    Если поле не входит в FIELD_NAME_MAP — возвращает оригинальное имя
    (это позволяет добавлять новые столбцы в Excel без изменения кода).
    """
    if not raw_name:
        return ""
    name = raw_name.strip()
    return FIELD_NAME_MAP.get(name, name)


def _is_http_url(s: str) -> bool:
    """Возвращает True, если строка начинается с http:// или https://.

    Это ЕДИНСТВЕННЫЙ допустимый формат значения для поля Image.
    Любые другие значения (локальные пути, относительные пути) будут
    проигнорированы фронтендом PWA через функцию isWorkingUrl().
    """
    return s.strip().lower().startswith(("http://", "https://"))


def _normalize_image_field(raw_value: str) -> str:
    """Нормализует значение поля Image из Excel.

    Единый принцип: только HTTP/HTTPS-ссылки.
    - Если значение — URL (http/https) → пропускаем как есть.
    - Если значение — что-то другое (локальный путь Windows,
      относительный путь, имя файла) → возвращаем пустую строку,
      чтобы в JSON не попал неработающий путь.
    """
    if not raw_value or not raw_value.strip():
        return ""
    if _is_http_url(raw_value):
        return raw_value.strip()
    print(f"  [ВНИМАНИЕ] Поле Image содержит не-URL значение, "
          f"оно будет проигнорировано в PWA: {raw_value[:80]}",
          file=sys.stderr)
    return ""


def _normalize_file_field(raw_value: str) -> str:
    """Нормализует значение поля Файл (ссылка на файл литературы).

    Аналогично полю Image: принимаются только HTTP/HTTPS-ссылки.
    """
    if not raw_value or not raw_value.strip():
        return ""
    if _is_http_url(raw_value):
        return raw_value.strip()
    return ""


# ============================================================
# Скачивание XLSX напрямую из Google Sheets
# (по образцу scripts/sync-projects.py)
# ============================================================
def download_file(spreadsheet_id, gid=None):
    """
    Скачивает XLSX-экспорт Google Sheets.

    URL: https://docs.google.com/spreadsheets/d/<ID>/export?format=xlsx[&gid=<GID>]
    Если gid не задан — экспортируется вся книга (все листы).
    """
    url = f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx'
    if gid:
        url += f'&gid={gid}'

    log(f'Скачивание: {url[:100]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120, allow_redirects=True)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code} — {resp.text[:200]}')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}. '
            f'Возможно, таблица не опубликована или нет доступа.'
        )

    filename = 'exam-tickets.xlsx'
    local_path = DOWNLOAD_DIR / filename
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')
    return local_path


# ============================================================
# Парсинг XLSX → JSON
# ============================================================
def convert_xlsx_to_json(xlsx_path, json_path):
    """Конвертирует xlsx в JSON.

    Структура вывода (для обратной совместимости с фронтендом):
    {
      "tickets-4": {
        "title": "Билеты на 4 разряд",
        "sheet": "4 разряд",
        "headers": ["id", "ticket_number", ...],
        "rows": [ {...}, {...} ],
        "total": N
      },
      "tickets-5": {...},
      "tickets-6": {...},
      "tickets-1000v": {...}
    }
    """
    log(f'Парсинг {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    log(f'Листы в файле: {wb.sheetnames}')

    all_data = {}
    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            log(f'Лист «{sheet_name}» не найден, пропускаю')
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else "" for h in rows[0]]
        normalized_headers = [_normalize_field_name(h) for h in headers]
        data_rows = []

        for row in rows[1:]:
            obj = {}
            for i, val in enumerate(row):
                if i < len(normalized_headers):
                    field_name = normalized_headers[i]

                    # Для числовых полей: если значение — datetime/time
                    # (бывает при формате Date/Time в Google Sheets),
                    # восстанавливаем исходное число (Excel serial number).
                    if field_name in NUMERIC_FIELD_NAMES and val is not None:
                        serial = datetime_to_serial(val)
                        if serial is not None:
                            obj[field_name] = format_serial_as_string(serial)
                            continue

                    cell_val = str(val) if val is not None else ""
                    # Единый принцип: только HTTP/HTTPS-ссылки.
                    if field_name == "image_url":
                        cell_val = _normalize_image_field(cell_val)
                    elif field_name == "file_url":
                        cell_val = _normalize_file_field(cell_val)
                    obj[field_name] = cell_val
            data_rows.append(obj)

        all_data[config["id"]] = {
            "title": config["title"],
            "sheet": sheet_name,
            "headers": normalized_headers,
            "rows": data_rows,
            "total": len(data_rows),
        }
        log(f'  {sheet_name}: {len(data_rows)} строк')

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    size = json_path.stat().st_size
    log(f'JSON сохранён: {json_path} ({size / 1024:.1f} КБ)')
    return True


def main():
    spreadsheet_id = os.environ.get('EXAM_TICKETS_SPREADSHEET_ID', '').strip() or DEFAULT_SPREADSHEET_ID
    gid = os.environ.get('EXAM_TICKETS_GID', '').strip() or None

    try:
        # 1. Скачать XLSX из Google Sheets
        local_file = download_file(spreadsheet_id, gid=gid)

        # 2. Конвертировать в JSON
        if not convert_xlsx_to_json(local_file, JSON_OUT):
            return 1
        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())
