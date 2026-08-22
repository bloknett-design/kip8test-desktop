#!/usr/bin/env python3
"""
Синхронизация перечня приборов КИП ИОС с Google Sheets.

Источник: https://docs.google.com/spreadsheets/d/1eUUwwulUvKUGWTgQ__XP-y7z1aEkt5Wy/edit
          (файл «Перечень КИП ИОС рабочий.xlsx», импортированный в Google Sheets)
Лист: "Приборы_app"

Скрипт:
1. Скачивает XLSX-экспорт напрямую из Google Sheets через export?format=xlsx.
   Google отдаёт файл без OAuth, если таблица доступна «у кого есть ссылка».
2. Парсит лист "Приборы_app" — заголовки в 1-й строке, данные со 2-й.
3. Сохраняет результат в data/devices.json.

Поле «Изображение» содержит прямые HTTPS-ссылки (Google Drive share-ссылки
вида https://drive.google.com/file/d/FILE_ID/view?usp=...). Они сохраняются
в JSON как есть — фронтенд PWA (index.html) сам конвертирует их в прямые
URL картинок через функцию gdriveShareToDirect().

Переменные окружения:
  DEVICES_SPREADSHEET_ID — ID Google Sheets
      (по умолчанию 1eUUwwulUvKUGWTgQ__XP-y7z1aEkt5Wy)
  DEVICES_SHEET_NAME — имя листа (по умолчанию "Приборы_app")
  DEVICES_GID — numeric ID листа (опционально; если задан, экспортирует
      конкретный лист через &gid=...). Если не задан — экспортируется вся книга.

Если нет интернета — используется уже существующий data/devices.json как заглушка.
"""
import os
import sys
import json
import re
from pathlib import Path
from datetime import datetime, time as dtime, date as ddate

import requests
import openpyxl


# ============================================================
# Настройки Google Sheets
# ============================================================
DEFAULT_SPREADSHEET_ID = '1eUUwwulUvKUGWTgQ__XP-y7z1aEkt5Wy'
DEFAULT_SHEET_NAME = 'Приборы_app'

DOWNLOAD_DIR = Path('/tmp/devices_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'devices.json'


def log(msg):
    print(f'[devices] {msg}', flush=True)


# ============================================================
# Восстановление числового значения из «даты/времени»
# ============================================================
# В Google Sheets числовые колонки иногда имеют формат Date/Time.
# openpyxl с data_only=True возвращает datetime/time вместо числа:
#   число 1   → datetime(1900, 1, 1)        (1900-01-01 = serial 1)
#   число 0.6 → time(14, 24)                (60% суток = 14:24)
#   число 62  → datetime(1900, 3, 3)        (62-й день от 1900-01-01)
# Решение: конвертируем datetime/time обратно в Excel serial number.
# Эпоха 1899-12-30 = 1900 date system (как в Google Sheets и Excel).
DATE_EPOCH = datetime(1899, 12, 30)

def datetime_to_serial(val):
    """Конвертирует datetime/time/date в Excel serial number (float).
    Возвращает None, если конвертация неприменима.
    """
    if isinstance(val, datetime):
        delta = val - DATE_EPOCH
        return round(delta.total_seconds() / 86400.0, 6)
    if isinstance(val, dtime):
        secs = val.hour * 3600 + val.minute * 60 + val.second + val.microsecond / 1e6
        return round(secs / 86400.0, 6)
    if isinstance(val, ddate):
        delta = datetime(val.year, val.month, val.day) - DATE_EPOCH
        return round(delta.total_seconds() / 86400.0, 6)
    return None


def format_serial_as_string(serial):
    """Форматирует serial number: int если целое, иначе trimmed float."""
    if abs(serial - round(serial)) < 1e-9:
        return str(int(round(serial)))
    return f'{serial:.4f}'.rstrip('0').rstrip('.')


# ============================================================
# Скачивание XLSX напрямую из Google Sheets
# ============================================================
def download_file(spreadsheet_id, gid=None):
    """
    Скачивает XLSX-экспорт Google Sheets.

    URL: https://docs.google.com/spreadsheets/d/<ID>/export?format=xlsx[&gid=<GID>]
    Если gid не задан — экспортируется вся книга (все листы).
    """
    url = f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx'
    if gid:
        url += f'&gid={gid}'

    log(f'Скачивание: {url[:100]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120, allow_redirects=True)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code} — {resp.text[:200]}')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}. '
            f'Возможно, таблица не опубликована или нет доступа.'
        )

    filename = 'devices.xlsx'
    local_path = DOWNLOAD_DIR / filename
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')
    return local_path


def parse_devices(xlsx_path, sheet_name):
    """
    Парсит лист sheet_name из XLSX-файла.
    Заголовки — в 1-й строке, данные — начиная со 2-й.
    Пропускает строки без ID или без Наименования.
    """
    log(f'Парсинг листа "{sheet_name}" из {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f'Лист "{sheet_name}" не найден. Доступные листы: {wb.sheetnames}')

    ws = wb[sheet_name]
    log(f'Размер листа: {ws.max_row} строк × {ws.max_column} колонок')

    # Читаем заголовки из 1-й строки
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ''
        headers.append(val)
    log(f'Заголовки ({len(headers)}): {headers}')

    # Читаем данные
    # Определяем, какие колонки должны быть числами (по заголовку).
    # В Google Sheets эти колонки иногда имеют формат Date/Time по ошибке,
    # и openpyxl возвращает datetime/time вместо числа. В таком случае
    # конвертируем datetime/time обратно в Excel serial number.
    NUMERIC_HEADERS_HINTS = ('ID', '№ прибора', '№ СБС', '№ САР', '№ проекта',
                             'Период ремонта', 'Позиция')

    def is_numeric_header(h):
        if not h:
            return False
        for hint in NUMERIC_HEADERS_HINTS:
            if hint in h:
                return True
        return False

    numeric_cols = set()
    for idx, h in enumerate(headers, 1):
        if is_numeric_header(h):
            numeric_cols.add(idx)
    if numeric_cols:
        log(f'Числовые колонки (по заголовку): {sorted(numeric_cols)} '
            f'→ {[headers[i-1] for i in sorted(numeric_cols)]}')

    devices = []
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value

            # Если колонка должна быть числовой, но значение — datetime/time,
            # восстанавливаем исходное число (Excel serial number).
            if col_idx in numeric_cols and val is not None:
                serial = datetime_to_serial(val)
                if serial is not None:
                    val = format_serial_as_string(serial)
                    row_values.append(val)
                    continue

            # Обработка дат для НЕчисловых колонок
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif val is not None:
                val = str(val).strip()
                # Убираем мягкие переносы и нормализуем пробелы
                val = val.replace('\xad', '').replace('\u00a0', ' ')
                val = re.sub(r'\s+', ' ', val).strip()
            else:
                val = ''
            row_values.append(val)

        # Создаём словарь "заголовок → значение"
        record = {}
        for h, v in zip(headers, row_values):
            if h:  # пропускаем пустые заголовки
                record[h] = v

        # Пропускаем строки без ID или без Наименования
        id_val = record.get('ID', '').strip()
        name_val = record.get('Наименование', '').strip()
        if not id_val and not name_val:
            skipped += 1
            continue

        # Если ID — число, преобразуем
        if id_val and id_val.isdigit():
            record['ID'] = int(id_val)

        devices.append(record)

    log(f'Распарсено записей: {len(devices)}, пропущено: {skipped}')
    return devices, headers


def main():
    spreadsheet_id = os.environ.get('DEVICES_SPREADSHEET_ID', '').strip() or DEFAULT_SPREADSHEET_ID
    sheet_name = os.environ.get('DEVICES_SHEET_NAME', '').strip() or DEFAULT_SHEET_NAME
    gid = os.environ.get('DEVICES_GID', '').strip() or None

    try:
        # 1. Скачать XLSX из Google Sheets
        local_file = download_file(spreadsheet_id, gid=gid)

        # 2. Распарсить лист
        devices, headers = parse_devices(local_file, sheet_name)

        # 3. Сохранить JSON (URL картинок проходят как есть — фронтенд
        #    сам конвертирует Google Drive share-ссылки в прямые URL)
        out = {
            'title': 'Перечень приборов КИП ИОС',
            'source': f'Google Sheets: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit',
            'sheet': sheet_name,
            'total_devices': len(devices),
            'headers': headers,
            'devices': devices,
        }
        JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
        with open(JSON_OUT, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        log(f'JSON сохранён: {JSON_OUT}')
        log(f'Всего приборов: {len(devices)}')

        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())
