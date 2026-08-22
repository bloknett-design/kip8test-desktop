#!/usr/bin/env python3
"""
Синхронизация расходомеров хозрасчётных с Google Sheets.

Источник: https://docs.google.com/spreadsheets/d/1enZSq7K8pwJVzaAI_tbXZtvATqARTxH0lSU4c-wc1eY/edit
          (файл «hozraschet_meters.gsheet»)
Лист: "hozraschet_meters"

Скрипт работает по тому же принципу, что и scripts/sync-cables.py
(раздел «Кабельный журнал»):

  1. Скачивает XLSX-экспорт напрямую из Google Sheets через export?format=xlsx.
     Google отдаёт файл без OAuth, если таблица доступна «у кого есть ссылка».
  2. Парсит лист "hozraschet_meters" — заголовки в 1-й строке, данные со 2-й.
     Структура столбцов (A–N):
       A: id         — номер позиции (1–12)
       B: hoz        — название (Хозрасчёт №1)
       C: param      — параметр (Расход пара в корпус 114)
       D: datePrev   — предыдущая дата
       E: dateCurr   — текущая дата
       F: prev       — предыдущие показания (число)
       G: curr       — текущие показания (число)
       H: unit       — единица измерения (т, м³)
       I: temp       — температура среды (число или пусто)
       J: Gcal       — гигакалории пара (число или пусто; только для расходомеров пара, Task 100)
       K: period     — периодичность (Ежедневно/Еженедельно/Ежемесячно)
       L: modRole    — роль пользователя, внёсшего последние изменения
       M: modName    — имя пользователя, внёсшего последние изменения
       N: modTimestamp — timestamp последнего ввода (Task 108 — для проверки «1 час на редактирование»)
  3. Сохраняет результат в data/flowmeters.json.

Переменные окружения:
  FLOWMETERS_SPREADSHEET_ID — ID Google Sheets
      (по умолчанию 1enZSq7K8pwJVzaAI_tbXZtvATqARTxH0lSU4c-wc1eY)
  FLOWMETERS_SHEET_NAME — имя листа (по умолчанию "hozraschet_meters")
  FLOWMETERS_GID — numeric ID листа (опционально; если задан, экспортирует
      конкретный лист через &gid=...). Если не задан — экспортируется вся книга.

Секреты НЕ требуются — таблица доступна «у кого есть ссылка»,
Google отдаёт XLSX через export?format=xlsx без OAuth.

Если нет интернета или API недоступен — используется уже существующий
data/flowmeters.json как заглушка (PWA продолжает работать с последними
закоммиченными данными).
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime, date as ddate

import requests
import openpyxl


# ============================================================
# Настройки Google Sheets
# ============================================================
DEFAULT_SPREADSHEET_ID = '1enZSq7K8pwJVzaAI_tbXZtvATqARTxH0lSU4c-wc1eY'
DEFAULT_SHEET_NAME = 'hozraschet_meters'

DOWNLOAD_DIR = Path('/tmp/flowmeters_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'flowmeters.json'


def log(msg):
    print(f'[flowmeters] {msg}', flush=True)


# ============================================================
# Скачивание XLSX напрямую из Google Sheets
# (по образцу scripts/sync-cables.py)
# ============================================================
def download_file(spreadsheet_id, gid=None):
    """
    Скачивает XLSX-экспорт Google Sheets.

    URL: https://docs.google.com/spreadsheets/d/<ID>/export?format=xlsx[&gid=<GID>]
    Если gid не задан — экспортируется вся книга (все листы).
    """
    url = f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx'
    if gid:
        url += f'&gid={gid}'

    log(f'Скачивание: {url[:100]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120, allow_redirects=True)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code} — {resp.text[:200]}')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}. '
            f'Возможно, таблица не опубликована или нет доступа.'
        )

    filename = 'flowmeters.xlsx'
    local_path = DOWNLOAD_DIR / filename
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')
    return local_path


# ============================================================
# Конвертация дат: Date object / DD.MM.YYYY → M/D/YYYY
# (тот же формат, что и сервер Flowmeter.gs → клиент)
# ============================================================
def sheet_to_client_date(val):
    """
    Конвертирует значение даты из Google Sheets в формат клиента M/D/YYYY.
    - datetime/date → M/D/YYYY
    - строка DD.MM.YYYY → M/D/YYYY
    - строка M/D/YYYY → как есть
    """
    if val is None:
        return ''
    if isinstance(val, (datetime, ddate)):
        return f'{val.month}/{val.day}/{val.year}'
    s = str(val).strip()
    if not s:
        return ''
    # DD.MM.YYYY → M/D/YYYY
    parts = s.split('.')
    if len(parts) == 3:
        try:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if 1 <= d <= 31 and 1 <= m <= 12 and y > 1900:
                return f'{m}/{d}/{y}'
        except ValueError:
            pass
    # M/D/YYYY — вернуть как есть
    if '/' in s:
        return s
    return s


# ============================================================
# Парсинг листа hozraschet_meters → data/flowmeters.json
# ============================================================
def parse_flowmeters(xlsx_path, sheet_name):
    """
    Парсит лист sheet_name из XLSX-файла.
    Заголовки — в 1-й строке, данные — начиная со 2-й.
    Структура жёсткая: 10 столбцов A–J (id, hoz, param, datePrev, dateCurr,
    prev, curr, unit, temp, period).
    """
    log(f'Парсинг листа "{sheet_name}" из {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Fallback на первый лист, если имя не найдено
    if sheet_name not in wb.sheetnames:
        log(f'Лист "{sheet_name}" не найден. Доступные: {wb.sheetnames}')
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            log(f'Используем первый лист: "{sheet_name}"')
        else:
            raise RuntimeError('В книге нет листов')

    ws = wb[sheet_name]
    log(f'Размер листа: {ws.max_row} строк × {ws.max_column} колонок')

    # Читаем заголовки из 1-й строки
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ''
        headers.append(val)
    # Убираем trailing пустые заголовки
    while headers and headers[-1] == '':
        headers.pop()
    log(f'Заголовки ({len(headers)}): {headers}')

    # Имена полей (маппинг столбцов → ключи клиента).
    # Строго соответствует заголовкам Google Sheets (A–N):
    #   A id, B hoz, C param, D datePrev, E dateCurr, F prev, G curr,
    #   H unit, I temp, J Gcal (Task 100), K period, L modRole, M modName,
    #   N modTimestamp (Task 108 — timestamp последнего ввода, для редактирования)
    # ВНИМАНИЕ: modRole идёт ПЕРЕД modName (как в заголовках таблицы).
    # В Task 100 они были перепутаны — это приводило к тому, что в flowmeters.json
    # modRole получал значение modName и наоборот (Task 101 — фикс).
    FIELD_NAMES = ['id', 'hoz', 'param', 'datePrev', 'dateCurr',
                   'prev', 'curr', 'unit', 'temp', 'gcal',
                   'period', 'modRole', 'modName', 'modTimestamp']

    meters = []
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        # Читаем 14 колонок (A–N)
        row_values = []
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            row_values.append(val)

        # Пропускаем пустые строки (нет id и нет hoz)
        if row_values[0] is None and row_values[1] is None:
            skipped += 1
            continue
        if str(row_values[0]).strip() == '' and str(row_values[1]).strip() == '':
            skipped += 1
            continue

        # Собираем meter-объект
        meter = {}
        for i, field in enumerate(FIELD_NAMES):
            val = row_values[i] if i < len(row_values) else None

            if field == 'id':
                # id — целое число
                if val is not None:
                    try:
                        meter['id'] = int(float(val))
                    except (ValueError, TypeError):
                        meter['id'] = i + 1
                else:
                    meter['id'] = i + 1

            elif field in ('datePrev', 'dateCurr'):
                # Даты: конвертируем в M/D/YYYY
                meter[field] = sheet_to_client_date(val)

            elif field in ('prev', 'curr'):
                # Числовые показания
                if val is not None:
                    try:
                        meter[field] = float(val)
                    except (ValueError, TypeError):
                        meter[field] = 0
                else:
                    meter[field] = 0

            elif field == 'temp':
                # Температура: число или null
                if val is not None and str(val).strip() != '':
                    try:
                        meter[field] = float(val)
                    except (ValueError, TypeError):
                        meter[field] = None
                else:
                    meter[field] = None

            elif field == 'gcal':
                # Гигакалории пара (Task 100): число или null
                # Заполняется только для расходомеров пара; для остальных = null
                if val is not None and str(val).strip() != '':
                    try:
                        meter[field] = float(val)
                    except (ValueError, TypeError):
                        meter[field] = None
                else:
                    meter[field] = None

            elif field == 'modTimestamp':
                # Timestamp последнего ввода (Task 108): ISO-строка или null
                # Используется для проверки «1 час на редактирование»
                if val is not None and isinstance(val, (datetime, ddate)):
                    meter[field] = val.isoformat()
                elif val is not None and str(val).strip() != '':
                    meter[field] = str(val).strip()
                else:
                    meter[field] = None

            else:
                # Строковые поля: hoz, param, unit, period
                meter[field] = str(val).strip() if val is not None else ''

        meters.append(meter)

    log(f'Распарсено позиций: {len(meters)}, пропущено: {skipped}')
    return meters, headers


def main():
    spreadsheet_id = os.environ.get('FLOWMETERS_SPREADSHEET_ID', '').strip() or DEFAULT_SPREADSHEET_ID
    sheet_name = os.environ.get('FLOWMETERS_SHEET_NAME', '').strip() or DEFAULT_SHEET_NAME
    gid = os.environ.get('FLOWMETERS_GID', '').strip() or None

    try:
        # 1. Скачать XLSX из Google Sheets
        local_file = download_file(spreadsheet_id, gid=gid)

        # 2. Распарсить лист
        meters, headers = parse_flowmeters(local_file, sheet_name)

        # 3. Сохранить JSON
        out = {
            'title': 'Расходомеры хозрасчётные',
            'source': f'Google Sheets: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit',
            'sheet': sheet_name,
            'total_meters': len(meters),
            'headers': headers,
            'meters': meters,
        }
        JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
        with open(JSON_OUT, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        log(f'JSON сохранён: {JSON_OUT}')
        log(f'Всего позиций: {len(meters)}')

        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())
